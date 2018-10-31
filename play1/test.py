import openpyxl
from openpyxl import Workbook, load_workbook
wb = load_workbook('cSpace Booking.xlsx')
#ws1 = wb["2018-07"]
#print (ws1['A2'].value)

ws1 = wb["Clients"]
#print (ws1['A2'].value)
rowHead = ws1['1']
nameC = None
issuesC = None

for cell in rowHead:
	if cell.value == "Name":
		nameC = cell.column
	if cell.value == "Issues":
		issuesC = cell.column



namecol = ws1[nameC]
issuescol=ws1[issuesC]


#fnamecol = []
#lnamecol = []
dictArray = []
#for col in ws1.iter_cols(min_row=2, max_col=1, max_row=21):
for i in range(1, len(namecol)): #namecol[1:]:
	sp = namecol[i].value.split()
	client = {
	"first": sp[0],
	"last": sp[1],
	"issues": issuescol[i].value 
	}
	dictArray.append(client)
	#fnamecol.append(sp[0])
	#lnamecol.append(sp[1])
	#print (namecol[-1].value)
#iss=ws1[issuesC]	
#issuecol = []
#for col in ws1.iter_cols(min_row=2, min_col=6, max_col=6, max_row=21):
#for cell in iss[1:]:
	
	#issuecol.append(cell)
	#print (issuecol[-1].value)
length = len(dictArray)
#print('First , Last | Issues')
html = """ <h2> Clients and Issues </h2>
			<table border = '1'>
				<tr>
					<td>
						First name
					</td>
					<td>
						Last name
					</td>
					<td>
						Issue
					</td>
				</tr>	
					"""
for i in range(0,length):
	html = html + "<tr>"
	html = html + "<td>" + dictArray[i]["first"] + "</td>"
	html = html + "<td>" + dictArray[i]["last"] + "</td>"
	html = html + "<td>" + str(dictArray[i]["issues"]) + "</td>"
	html = html + "</tr>"

html = html + """</table>
					</h2>"""
	#print (fnamecol[i],",",lnamecol[i], " | ",  issuecol[i].value)
#print (fnamecol[0])

w = open('clients.html' , 'w')
w.write(html)
w.close()